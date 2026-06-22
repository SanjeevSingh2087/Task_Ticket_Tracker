from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io
import os
import copy

app = Flask(__name__)
CORS(app)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "HO_Ticket_Tracker.xlsx")
EXCEL_HEADERS = [
    "Ticket No", "Date", "Time Opened", "Opened By", "Vendor",
    "Task Detail", "Priority", "Assigned to", "Status", "Update", "Last Modified"
]


def ensure_excel_workbook():
    if not os.path.exists(EXCEL_PATH):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tasks"
        ws.append(EXCEL_HEADERS)
        wb.save(EXCEL_PATH)
    else:
        # Migrate: add any missing columns to existing workbook
        migrate_excel_columns()


def migrate_excel_columns():
    """Add any missing columns from EXCEL_HEADERS to the existing workbook."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    existing_headers = [cell.value.strip() if cell.value else '' for cell in ws[1]]
    changed = False
    for col_name in EXCEL_HEADERS:
        if col_name not in existing_headers:
            next_col = len(existing_headers) + 1
            ws.cell(row=1, column=next_col, value=col_name)
            existing_headers.append(col_name)
            changed = True
    # Populate Last Modified for existing rows if empty
    if "Last Modified" in existing_headers:
        last_mod_idx = existing_headers.index("Last Modified") + 1
        date_idx = existing_headers.index("Date") + 1 if "Date" in existing_headers else None
        time_idx = existing_headers.index("Time Opened") + 1 if "Time Opened" in existing_headers else None
        for row in ws.iter_rows(min_row=2):
            if not row[last_mod_idx - 1].value:
                date_val = row[date_idx - 1].value if date_idx else None
                time_val = row[time_idx - 1].value if time_idx else None
                if date_val and time_val:
                    last_mod = f"{date_val} {time_val}"
                elif date_val:
                    last_mod = str(date_val)
                else:
                    continue
                row[last_mod_idx - 1].value = last_mod
                changed = True
    if changed:
        wb.save(EXCEL_PATH)


def load_df():
    ensure_excel_workbook()
    df = pd.read_excel(EXCEL_PATH, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    df = df.fillna("")
    return df

def match_ticket(val, ticket_no):
    if val is None or ticket_no is None:
        return False
    v_str = str(val).strip()
    t_str = str(ticket_no).strip()
    if v_str == t_str:
        return True
    try:
        if float(v_str) == float(t_str):
            return True
    except ValueError:
        pass
    return False

def save_row_to_excel(row_data, ticket_no=None):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    header_map = {}
    for cell in ws[1]:
        if cell.value:
            header_map[cell.value.strip()] = cell.column

    # Determine if update or insert
    existing_row = None
    if ticket_no:
        for row in ws.iter_rows(min_row=2):
            val = row[0].value
            if match_ticket(val, ticket_no):
                existing_row = row[0].row
                break

    col_map = {
        "Ticket No": "Ticket No",
        "Date": "Date",
        "Time Opened": "Time Opened",
        "Opened By": "Opened By",
        "Vendor": "Vendor",
        "Task Detail": "Task Detail",
        "Priority": "Priority",
        "Assigned to": "Assigned to",
        "Status": "Status",
        "Update": "Update",
        "Last Modified": "Last Modified"
    }

    if existing_row:
        target_row = existing_row
    else:
        target_row = ws.max_row + 1
        # New ticket: assign ticket number
        max_ticket = 0
        for row in ws.iter_rows(min_row=2, max_col=1):
            for cell in row:
                try:
                    v = int(float(str(cell.value).strip()))
                    if v > max_ticket:
                        max_ticket = v
                except:
                    pass
        row_data["Ticket No"] = str(max_ticket + 1)

    for field, col_name in col_map.items():
        col_idx = header_map.get(col_name)
        if col_idx is None:
            # try stripped
            for k, v in header_map.items():
                if k.strip() == col_name.strip():
                    col_idx = v
                    break
        if col_idx and field in row_data:
            ws.cell(row=target_row, column=col_idx, value=row_data[field])

    wb.save(EXCEL_PATH)
    return row_data.get("Ticket No")

@app.route("/")
def index():
    return send_file(os.path.join(os.path.dirname(__file__), "dashboard.html"))

@app.route("/api/tasks", methods=["GET"])
def get_tasks():
    df = load_df()
    tasks = df.to_dict(orient="records")
    # Rename "Assigned to " -> "Assigned to"
    cleaned = []
    for t in tasks:
        ct = {}
        for k, v in t.items():
            ct[k.strip()] = v
        cleaned.append(ct)
    return jsonify(cleaned)

@app.route("/api/tasks", methods=["POST"])
def add_task():
    data = request.json
    now = datetime.now()
    data["Date"] = now.strftime("%Y-%m-%d")
    data["Time Opened"] = now.strftime("%H:%M:%S")
    data["Last Modified"] = now.strftime("%Y-%m-%d %H:%M:%S")
    ticket_no = save_row_to_excel(data)
    return jsonify({"success": True, "ticket_no": ticket_no})

@app.route("/api/tasks/<ticket_no>", methods=["PUT"])
def update_task(ticket_no):
    data = request.json
    data["Ticket No"] = ticket_no
    data["Last Modified"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    save_row_to_excel(data, ticket_no=ticket_no)
    return jsonify({"success": True})

@app.route("/api/verify-delete-code", methods=["POST"])
def verify_delete_code():
    data = request.json
    code = data.get("code", "").strip()
    DELETE_CODE = "1947"
    is_valid = code == DELETE_CODE
    return jsonify({"valid": is_valid})

@app.route("/api/tasks/<ticket_no>", methods=["DELETE"])
def delete_task(ticket_no):
    data = request.json or {}
    code = data.get("code", "").strip()
    DELETE_CODE = "1947"
    
    # Verify code before allowing deletion
    if code != DELETE_CODE:
        return jsonify({"success": False, "error": "Invalid code. Task cannot be deleted."}), 403
    
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if match_ticket(row[0].value, ticket_no):
            ws.delete_rows(row[0].row)
            break
    wb.save(EXCEL_PATH)
    return jsonify({"success": True})

@app.route("/api/download", methods=["GET"])
def download_filtered():
    status_filter = request.args.get("status", "open")
    dept_filter = request.args.get("dept")
    employee_filter = request.args.get("employee")
    priority_filter = request.args.get("priority")
    search_filter = request.args.get("search")
    
    df = load_df()
    df.columns = [c.strip() for c in df.columns]
    
    if status_filter == "open":
        filtered = df[df["Status"].str.contains("Open", case=False, na=False)]
    elif status_filter == "closed":
        filtered = df[~df["Status"].str.contains("Open", case=False, na=False)]
    else:
        filtered = df

    if dept_filter:
        filtered = filtered[filtered["Vendor"].str.contains(dept_filter, case=False, na=False)]

    if employee_filter:
        filtered = filtered[filtered["Assigned to"].str.contains(employee_filter, case=False, na=False)]
        
    if priority_filter:
        filtered = filtered[filtered["Priority"].str.contains(priority_filter, case=False, na=False)]
        
    if search_filter:
        # Advanced search: support #ticketnumber syntax
        if search_filter.startswith('#'):
            ticket_num = search_filter[1:].strip()
            filtered = filtered[filtered['Ticket No'].astype(str).str.strip() == ticket_num]
        else:
            search_lower = search_filter.lower()
            mask = filtered.apply(lambda row: row.astype(str).str.lower().str.contains(search_lower).any(), axis=1)
            filtered = filtered[mask]

    # Style output
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Filtered Tasks"

    headers = list(filtered.columns)
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    priority_fills = {
        "High": PatternFill("solid", fgColor="FEE2E2"),
        "Medium": PatternFill("solid", fgColor="FEF9C3"),
        "Low": PatternFill("solid", fgColor="DCFCE7"),
    }
    status_fonts = {
        "Open": Font(color="DC2626", bold=True, name="Calibri"),
        "Open – Partial done": Font(color="D97706", bold=True, name="Calibri"),
    }

    for ri, (_, row) in enumerate(filtered.iterrows(), 2):
        priority = str(row.get("Priority", ""))
        status = str(row.get("Status", ""))
        row_fill = priority_fills.get(priority, PatternFill("solid", fgColor="F8FAFC"))
        for ci, h in enumerate(headers, 1):
            val = row[h]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if h == "Status":
                for k, f in status_fonts.items():
                    if k.lower() in status.lower():
                        cell.font = f
                        break

    col_widths = {"Ticket No": 10, "Date": 14, "Time Opened": 12, "Opened By": 15,
                  "Vendor": 22, "Task Detail": 50, "Priority": 10,
                  "Assigned to": 25, "Status": 22, "Update": 30}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = col_widths.get(h, 18)

    ws.row_dimensions[1].height = 25
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # Generate filename based on filters
    filename_parts = ["Tasks"]
    if status_filter == "open":
        filename_parts.append("Open")
    if dept_filter:
        filename_parts.append(dept_filter.replace(" ", "_"))
    if employee_filter:
        filename_parts.append(employee_filter.replace(" ", "_"))

    fname = f"{'_'.join(filename_parts)}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True)

@app.route("/api/department/<dept_name>", methods=["GET"])
def get_department_tasks(dept_name):
    df = load_df()
    df.columns = [c.strip() for c in df.columns]
    dept_tasks = df[df["Vendor"].str.contains(dept_name, case=False, na=False)]

    # Department stats
    open_tasks = dept_tasks[dept_tasks["Status"].str.contains("Open", case=False, na=False)]
    closed_tasks = dept_tasks[~dept_tasks["Status"].str.contains("Open", case=False, na=False)]

    # Employee breakdown for this department
    employee_breakdown = open_tasks.groupby("Assigned to").size().reset_index(name="count")
    employee_breakdown = employee_breakdown.sort_values("count", ascending=False)

    return jsonify({
        "department": dept_name,
        "total_tasks": len(dept_tasks),
        "open_tasks": len(open_tasks),
        "closed_tasks": len(closed_tasks),
        "tasks": dept_tasks.to_dict(orient="records"),
        "employee_breakdown": employee_breakdown.to_dict(orient="records")
    })

@app.route("/api/employee/<employee_name>", methods=["GET"])
def get_employee_tasks(employee_name):
    df = load_df()
    df.columns = [c.strip() for c in df.columns]
    employee_tasks = df[df["Assigned to"].str.contains(employee_name, case=False, na=False)]

    # Employee stats
    open_tasks = employee_tasks[employee_tasks["Status"].str.contains("Open", case=False, na=False)]
    closed_tasks = employee_tasks[~employee_tasks["Status"].str.contains("Open", case=False, na=False)]

    return jsonify({
        "employee": employee_name,
        "total_tasks": len(employee_tasks),
        "open_tasks": len(open_tasks),
        "closed_tasks": len(closed_tasks),
        "tasks": employee_tasks.to_dict(orient="records")
    })

@app.route("/api/stats", methods=["GET"])
def get_stats():
    df = load_df()
    df.columns = [c.strip() for c in df.columns]
    total = len(df)
    open_tasks = df[df["Status"].str.contains("Open", case=False, na=False)]
    closed_tasks = df[~df["Status"].str.contains("Open", case=False, na=False)]
    high_open = open_tasks[open_tasks["Priority"].str.lower() == "high"]
    medium_open = open_tasks[open_tasks["Priority"].str.lower() == "medium"]

    dept_breakdown = open_tasks.groupby("Vendor").size().reset_index(name="count")
    dept_breakdown = dept_breakdown.sort_values("count", ascending=False).head(10)

    # Employee-wise breakdown
    employee_breakdown = open_tasks.groupby("Assigned to").size().reset_index(name="count")
    employee_breakdown = employee_breakdown.sort_values("count", ascending=False).head(10)

    priority_open = open_tasks.groupby("Priority").size().to_dict()

    # Weekly closed tasks tracking
    closed_tasks_copy = closed_tasks.copy()
    closed_tasks_copy["Date"] = pd.to_datetime(closed_tasks_copy["Date"], errors='coerce')
    weekly_closed = closed_tasks_copy.groupby(pd.Grouper(key="Date", freq="W")).size().reset_index(name="count")
    weekly_closed = weekly_closed.sort_values("Date", ascending=False).head(4)

    return jsonify({
        "total": total,
        "open": len(open_tasks),
        "closed": len(closed_tasks),
        "high_open": len(high_open),
        "medium_open": len(medium_open),
        "dept_breakdown": dept_breakdown.to_dict(orient="records"),
        "employee_breakdown": employee_breakdown.to_dict(orient="records"),
        "priority_breakdown": priority_open,
        "weekly_closed": weekly_closed.to_dict(orient="records")
    })

if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5055))
    app.run(debug=False, port=port, host="0.0.0.0")
